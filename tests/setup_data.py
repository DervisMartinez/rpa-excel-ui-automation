from pathlib import Path
from openpyxl import Workbook
import sys

def setup():
    base_dir = Path(__file__).parent.parent
    data_dir = base_dir / ".data"
    input_dir = data_dir / "input"
    output_dir = data_dir / "output"
    
    input_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    origen_path = input_dir / "origen.xlsx"
    
    if not origen_path.exists():
        print(f"Creando archivo de prueba en {origen_path}")
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Datos de prueba"
        ws["A2"] = "Para automatización RPA"
        wb.save(origen_path)
        print("Archivo creado con éxito.")
    else:
        print("El archivo origen ya existe.")

if __name__ == "__main__":
    setup()
